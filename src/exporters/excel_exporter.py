from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.domain.lead_report import LeadReport


class ExcelExporterError(RuntimeError):
    """Raised when an Excel export cannot be completed."""


class ExcelExporter:
    """Export a lead report to a structured Excel workbook."""

    def export(self, report: LeadReport, output_path: str) -> None:
        """Export the supplied lead report to an Excel file."""
        try:
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            workbook = Workbook()
            workbook.remove(workbook.active)

            self._create_summary_sheet(workbook, report)
            self._create_analysis_sheet(workbook, report)
            self._create_recommendation_sheet(workbook, report)

            workbook.save(output_file)
        except Exception as exc:  # pragma: no cover - defensive boundary
            raise ExcelExporterError(f"Failed to export Excel report: {exc}") from exc

    def _create_summary_sheet(self, workbook: Workbook, report: LeadReport) -> None:
        """Create the executive summary worksheet."""
        sheet = workbook.create_sheet("Executive Summary")
        rows: list[tuple[str, Any]] = [
            ("Company Name", report.company.name),
            ("Industry", report.company.industry or "N/A"),
            ("Headquarters", report.company.headquarters or "N/A"),
            ("Digital Maturity", report.analysis.digital_maturity),
            ("Recommended ERP", report.recommendation.recommended_erp),
            ("Confidence Score", report.analysis.confidence_score),
            ("Generated At", report.generated_at.isoformat()),
        ]
        self._write_rows(sheet, rows)

    def _create_analysis_sheet(self, workbook: Workbook, report: LeadReport) -> None:
        """Create the business analysis worksheet."""
        sheet = workbook.create_sheet("Business Analysis")
        rows: list[tuple[str, Any]] = [
            ("Business Challenges", self._join_items(report.analysis.business_challenges)),
            ("ERP Opportunities", self._join_items(report.analysis.erp_opportunities)),
            ("Transformation Goals", self._join_items(report.analysis.transformation_goals)),
            ("Decision Makers", self._join_items(report.analysis.decision_makers)),
        ]
        self._write_rows(sheet, rows)

    def _create_recommendation_sheet(self, workbook: Workbook, report: LeadReport) -> None:
        """Create the ERP recommendation worksheet."""
        sheet = workbook.create_sheet("ERP Recommendation")
        rows: list[tuple[str, Any]] = [
            ("Recommended ERP", report.recommendation.recommended_erp),
            ("Recommended Modules", self._join_items(report.recommendation.recommended_modules)),
            ("Business Justification", report.recommendation.rationale),
            ("Expected Benefits", self._join_items(report.recommendation.expected_business_value)),
            ("Implementation Risks", self._join_items(report.recommendation.implementation_risks)),
            ("Priority", report.recommendation.implementation_priority),
            ("Next Sales Actions", self._join_items(report.recommendation.expected_business_value)),
        ]
        self._write_rows(sheet, rows)

    def _write_rows(self, sheet: Any, rows: list[tuple[str, Any]]) -> None:
        """Write rows and style headers for a worksheet."""
        for row_index, (label, value) in enumerate(rows, start=1):
            sheet.cell(row=row_index, column=1, value=label)
            sheet.cell(row=row_index, column=2, value=value)

        header_font = Font(bold=True)
        for cell in sheet["A1:A" + str(len(rows))]:
            for item in cell:
                item.font = header_font

        self._auto_size_columns(sheet)

    def _auto_size_columns(self, sheet: Any) -> None:
        """Auto-size all worksheet columns."""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def _join_items(self, items: list[str]) -> str:
        """Join a list of strings into a single cell value."""
        return "\n".join(items) if items else "N/A"
